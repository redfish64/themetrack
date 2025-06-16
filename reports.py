from dataclasses import dataclass
from enum import Enum, auto
from functools import partial
import re
from typing import Any, Callable
import pandas as pd
import ftypes
from openpyxl.styles import Font
import array_log as al
import util
from typing import Sequence

HOLDINGS_WS_TITLE = 'Holdings Input Data'
PICK_WS_TITLE = 'Pick Input Data'
JOINED_DATA_WS_TITLE = 'Joined Data'
NA_STR_NAME = '(none)'

header_font = Font(bold=True, italic=True)
normal_line_font = Font()
total_line_font = Font(bold=True)

def style_simple_report_ws(ws, num_lines):
    #Make header stylish
    for cell in ws[1]:
        cell.font = header_font

    #if we don't set to anything, will default black color, which is not readable in dark mode in libreoffice
    for row_idx in range(2,num_lines+2):
        for cell in ws[row_idx]:
            cell.font = normal_line_font


def style_report_ws(ws, num_lines):
    #Make header stylish
    for cell in ws[1]:
        cell.font = header_font

    #if we don't set to anything, will default black color, which is not readable in dark mode in libreoffice
    for row_idx in range(2,num_lines+1):
        for cell in ws[row_idx]:
            cell.font = normal_line_font

    #Make last line (the total) stylish
    for cell in ws[num_lines+1]:
        cell.font = total_line_font



def calc_max_len(format_code, value):
    # Extract prefix from the format code
    prefix_match = re.match(r'\"([^\"]+)\"', format_code)
    prefix = prefix_match.group(1) if prefix_match else ''

    # Check if the format is for percentage
    if isinstance(value,(int, float)):
        if re.match(r'0\.0+%', format_code):
            # Percentage format
            decimal_places = format_code.count('0') - 2
            formatted_value = f"{value * 100:.{decimal_places}f}%"
        elif re.match(r'\"\D*\"#,##0\.00', format_code):
            # Currency format with any prefix
            formatted_value = f"{prefix}{value:,.2f}"
        else:
            # General numeric format
            decimal_places = format_code.count('0')
            formatted_value = f"{value:,.{decimal_places}f}"
    else: #format_code == '@':
        # General text format
        formatted_value = str(value)

        #TODO 3 this function isn't perfect, but it handles common formats. Otherwise it falls back to this generic clause where we add 2 for some slop
        return len(formatted_value) + 2 
    return len(formatted_value)

def update_number_formats(excel_formats,ws,num_rows,native_currency_code):
    """Updates the style of the cell according to the column type, so percentages look like %0.01,
        money looks like $1,000,000,000.00 etc.

        Also does a somewhat sloppy job of calcuating column width. Can't really do better without
        actually talking to excel

    Args:
        excel_formats (_type_): excel format for each row
        ws (_type_): worksheet to modify
        num_rows (_type_): number of rows to modify (after the header)
    """
    for (excel_format,col) in zip(excel_formats,ws.iter_cols()):
        max_len = len(col[0].value)
        for cell in col[1:num_rows+1]:
            cell.number_format = excel_format
            max_len = max(calc_max_len(excel_format, cell.value),max_len)

        adjusted_width = (max_len + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width      


def calc_performance_gains_for_cat(config, cat_column, res_df, joined_df):
    for period in config.hist_perf_periods:
        # ---- column names -----------------------------------------------------------
        qty_col         = ftypes.SpecialColumns.RQuantity.get_col_name()
        start_price_col = f"{ftypes.ADJ_CLOSE_START_PRICE_PREFIX}{period}"
        end_price_col   = f"{ftypes.ADJ_CLOSE_END_PRICE_PREFIX}{period}"
        perf_col        = f"{ftypes.GAIN_LOSS_PREFIX}{period}"                       # % perf
        flag_col        = f"{ftypes.GAIN_LOSS_NOT_ALL_DATA_PRESENT_PREFIX}{period}"

        required = [cat_column, qty_col, start_price_col, end_price_col]
        missing_cols = [c for c in required if c not in joined_df.columns]
        if missing_cols:
            raise KeyError(f"joined_df is missing required columns: {missing_cols}")

        # ---- completeness masks -----------------------------------------------------
        complete_mask   = joined_df[[qty_col, start_price_col, end_price_col]].notna().all(axis=1)
        incomplete_mask = ~complete_mask

        # ---- per-row numbers --------------------------------------------------------
        per_row_gain  = (joined_df[end_price_col] - joined_df[start_price_col]) * joined_df[qty_col]
        per_row_gain  = per_row_gain.where(complete_mask, 0.0)                         # ignore bad rows

        per_row_basis = joined_df[start_price_col] * joined_df[qty_col]               # initial value
        per_row_basis = per_row_basis.where(complete_mask, 0.0)

        # ---- aggregate to category level -------------------------------------------
        cat_gain  = per_row_gain.groupby(joined_df[cat_column]).sum()
        cat_basis = per_row_basis.groupby(joined_df[cat_column]).sum()

        cat_pct_perf = ((cat_gain / cat_basis.replace(0, pd.NA))        # avoid /0 → inf
                        .fillna(0.0))                            # percentage

        cat_any_missing = incomplete_mask.groupby(joined_df[cat_column]).any()

        # ---- write into res_df ------------------------------------------------------
        res_df[perf_col] = res_df[cat_column].map(cat_pct_perf).fillna(0.0)            # % gain
        res_df[flag_col] = res_df[cat_column].map(
            lambda c: "*" if cat_any_missing.get(c, False) else ""
        )

def calc_performance_gains_for_stocks(config, res_df):
    for period in config.hist_perf_periods:
        # ---- column names -----------------------------------------------------------
        start_price_col = f"{ftypes.ADJ_CLOSE_START_PRICE_PREFIX}{period}"
        end_price_col   = f"{ftypes.ADJ_CLOSE_END_PRICE_PREFIX}{period}"
        perf_col        = f"{ftypes.GAIN_LOSS_PREFIX}{period}"                       # % perf

        # ---- per-row numbers --------------------------------------------------------
        res_df[perf_col] = (res_df[end_price_col] - res_df[start_price_col]) / res_df[start_price_col] - 1.0


def make_portfolio_report(config : ftypes.Config, report_config : ftypes.ReportConfig, joined_df : pd.DataFrame, holdings_df : pd.DataFrame, 
                           picks_df : pd.DataFrame, native_currency_code : str
                          ):

    total_sum = joined_df[ftypes.SpecialColumns.RCurrValue.get_col_name()].sum()

    def get_total_perc(row):
        val = row[ftypes.SpecialColumns.RCurrValue.get_col_name()]

        return val/total_sum
    
    category_df = pd.pivot_table(joined_df, values=ftypes.SpecialColumns.RCurrValue.get_col_name(), 
                                    index=[report_config.cat_column],
                                    aggfunc="sum", fill_value=0).copy()

    if(report_config.is_cat_type):
        res_df = category_df
        res_df.reset_index(names=[report_config.cat_column],inplace=True)
        res_df[ftypes.SpecialColumns.RTotalPerc.get_col_name()] = category_df.apply(get_total_perc,axis=1)
        calc_performance_gains_for_cat(config, report_config.cat_column, res_df, joined_df)
    else:
        category_df[ftypes.SpecialColumns.RCatTotalPerc.get_col_name()] = category_df.apply(get_total_perc,axis=1)
        res_df = joined_df.copy()
    
        def get_cat_perc(row):
            cat = row[report_config.cat_column]
            val = row[ftypes.SpecialColumns.RCurrValue.get_col_name()]
            cat_val = category_df.loc[cat,ftypes.SpecialColumns.RCurrValue.get_col_name()]

            return val/cat_val


        res_df[ftypes.SpecialColumns.RCatPerc.get_col_name()] = res_df.apply(get_cat_perc,axis=1)
        res_df[ftypes.SpecialColumns.RTotalPerc.get_col_name()] = res_df.apply(get_total_perc,axis=1)

        def row_filter_fn(bitmask):
            def res_fn(row):
                val = row[ftypes.SpecialColumns.RCurrValue.get_col_name()]
                if val != 0.0 and not pd.isna(val):
                    return True

                bm = util.default_df_val(row[ftypes.SpecialColumns.DJoinAllBitMask.get_col_name()],0)
                return (bm & bitmask) != 0
            
            return res_fn

        #we need to filter out empty rows that don't correspond to our report type. For capgains, we don't want empty divi
        #portfolio rows, since there are so many and its distracting. For divi, vic versa.
        res_df = res_df[res_df.apply(row_filter_fn(report_config.always_show_pick_bitmask),axis=1)]

        calc_performance_gains_for_stocks(config, res_df)

        # TODO 2.5 reenable this, using a different column per report or something...
        # def get_cat_total_perc_for_df(row):
        #     cat = row[report_config.cat_column]
        #     cat_val = category_df.loc[cat][ftypes.SpecialColumns.RCurrValue.get_col_name()]

        #     return cat_val/total_sum
        # we feed the data back into the main df, as well, so that the data page has all the data
        # joined_df[ftypes.SpecialColumns.RCatPerc.get_col_name()] = stocks_df[ftypes.SpecialColumns.RCatPerc.get_col_name()]
        # joined_df[ftypes.SpecialColumns.RTotalPerc.get_col_name()] = stocks_df[ftypes.SpecialColumns.RTotalPerc.get_col_name()]
        # joined_df[ftypes.SpecialColumns.RCatTotalPerc.get_col_name()] = joined_df.apply(get_cat_total_perc_for_df,axis=1)

    res_df.sort_values(by=report_config.column_order,inplace=True)

    res_df = res_df[[r[0] for r in report_config.columns]]

    res_totals = pd.DataFrame(res_df[report_config.sum_columns].sum()).T
    res_totals[res_df.columns[0]] = 'Total'

    res_df = pd.concat([res_df,res_totals],ignore_index=True)

    res_df.rename(columns={ name : display_as for name,display_as,excel_format in report_config.columns}, inplace=True)

    def final_step_fn(writer):
        res_df.to_excel(writer, index=False, sheet_name=report_config.name)
        ws = writer.sheets[report_config.name]
        style_report_ws(ws,res_df.shape[0])

        excel_formats = [r[2] for r in report_config.columns]
        update_number_formats(excel_formats,ws,res_df.shape[0],native_currency_code)

    return final_step_fn


def make_report_workbook(orig_joined_df : pd.DataFrame, holdings_df : pd.DataFrame, picks_df : pd.DataFrame, native_currency_code : str, 
                         rules_log : al.Log, config : ftypes.Config,output_file : str) -> pd.DataFrame:
    
    #PERF, this table is pretty large, but we don't want to go mucking with it here and then use it for something else later
    joined_df = orig_joined_df.copy()

    joined_df[ftypes.SpecialColumns.RTheme.get_col_name()] = joined_df[ftypes.SpecialColumns.RTheme.get_col_name()].fillna(NA_STR_NAME)
    joined_df[ftypes.SpecialColumns.RSector.get_col_name()] = joined_df[ftypes.SpecialColumns.RSector.get_col_name()].fillna(NA_STR_NAME)

    #make_portfolio_report returns a function here because debugging while within a "with pd.ExcelWriter..." is a pain,
    #so we minimize the amount of code inside it
    report_writer_fn_list = [make_portfolio_report(config, report_config, joined_df,holdings_df,picks_df,native_currency_code) for report_config in config.reports]
    
    # Export to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for rw_fn in report_writer_fn_list:
            rw_fn(writer)

        def add_df(res_df, title, writer):
            res_df.to_excel(writer, index=False, sheet_name=title)
            ws = writer.sheets[title]
            style_simple_report_ws(ws,res_df.shape[0])

        add_df(holdings_df, HOLDINGS_WS_TITLE, writer)
        add_df(picks_df, PICK_WS_TITLE, writer)
        add_df(orig_joined_df, JOINED_DATA_WS_TITLE, writer)
        





